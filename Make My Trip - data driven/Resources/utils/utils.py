from robot.api.deco import keyword
import openpyxl
import random
import calendar
from datetime import datetime
from datetime import timedelta

@keyword("Fetch Testdata By Id")  #can use this name to access this function in robot
def fetch_testdata_by_id(excel_file_path, sheet_name, tc_id):
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        # tc_id_column_index = 1
        tc_id_column_index = None

        for idx, header in enumerate(headers, start=1):
            if header and str(header).upper() in ['TC_ID', 'TEST_ID', '${TC_ID}', '${TEST_ID}']:
                tc_id_column_index = idx
                break
        if tc_id_column_index is None:
            raise ValueError(f"TC_ID column not found in Excel headers: {headers}")
        test_data = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if row[tc_id_column_index - 1].value == tc_id:
                for col_idx, header in enumerate(headers, start=1):
                    if header:
                        clean_header = str(header).replace('${', '').replace('}', '').strip()
                        test_data[clean_header] = row[col_idx - 1].value
                break
        if not test_data:
            raise ValueError(f"Test Case ID '{tc_id}' not found in Excel file")
        wb.close()
        return test_data
    except Exception as e:
        raise Exception(f"Error fetching test data for TC_ID '{tc_id}': {str(e)}")


@keyword("Create Dict With Dot Access")
def create_dict_with_dot_access(data_dict):
    """Convert dictionary to object with dot notation access for Robot Framework"""
    class DotDict:
        def __init__(self, d):
            for key, value in d.items():
                setattr(self, key, value)
        def __getitem__(self, key):
            return getattr(self, key)
        def __str__(self):
            return str(self.__dict__)
    return DotDict(data_dict)

@keyword
def get_page2_heading(frm, to):
    return f"{frm} to {to} Bus"

@keyword
def get_random_date(current_date_str):
    # Parse input date like "28 Jan26"
    current_date = datetime.strptime(current_date_str, "%d %b%y")

    # Calculate month + 3 (calendar-based)
    target_month = current_date.month + 3
    target_year = current_date.year

    if target_month > 12:
        target_month -= 12
        target_year += 1

    # Get last day of the target month
    last_day = calendar.monthrange(target_year, target_month)[1]
    max_date = datetime(target_year, target_month, last_day)

    # Total days in allowed range
    total_days = (max_date - current_date).days

    # Generate random offset (including current date)
    random_offset = random.randint(0, total_days)

    random_date = current_date + timedelta(days=random_offset)

    return random_date.strftime("%b %d %Y")
